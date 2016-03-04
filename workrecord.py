# -*- coding: utf-8 -*- 
# To covert the work record to text.
# By Jun Lv, 20114 Jun.Lv@nsn.com
# This software is placed into the private domain
# Revision date: Feb 25, 2016
# Version: 1.0

import argparse
import os
import datetime
import time
from  openpyxl.reader.excel  import  load_workbook 

def convert(excelfile,workdate):
    wb = load_workbook(excelfile )  
    ws = wb.get_sheet_by_name("workrecord") 
    lastrow=ws.max_row
    text1=u"PS工作内容"
    text2=u"次日值守:"
    flag=0
    rowno=2
    workdate=time.strptime(workdate,"%Y%m%d")  
    workdate=datetime.date(workdate[0],workdate[1],workdate[2])    
         
    while (rowno<lastrow  and flag==0):            
        if(ws.cell(row=rowno,column=1).value!=None):
            if ws.cell(row=rowno,column=1).value.date()==workdate:
                flag=1
            else:
                rowno=rowno+1
        else:
            rowno=rowno+1  
            
    if(rowno==lastrow):
                print "No result found!!!"
                return        
        
    print "%s%15s" % (workdate,text1)
    if("A"+str(rowno) not in ws.merged_cells):
        print  "%-55s%-40s" %(ws.cell(row=rowno,column=2).value,ws.cell(row=rowno,column=4).value)        
        print  "%s%s" %(text2,ws.cell(row=rowno,column=5).value)
    else:
        beginrow=rowno
        flag=1
        while(rowno<lastrow and flag==1):       
            print  "%-55s%-40s" %(ws.cell(row=rowno,column=2).value,ws.cell(row=rowno,column=4).value)
            rowno=rowno+1
            if(ws.cell(row=rowno,column=1).value!=None):
                flag=0
        print  "%s%s" %(text2,ws.cell(row=beginrow,column=5).value)


if __name__== '__main__':   
    # 参数处理说明
    parser = argparse.ArgumentParser(description='Reading work record from xlsx and convert to text')       
    parser.add_argument("-i", metavar="inputFile",default='workrecord.xlsx',help="work record name")  
    parser.add_argument("-d", metavar="date",default=None,help="work date")     
    parser.add_argument('-v','--version', action='version', version='%(prog)s 1.0')
    args = parser.parse_args()      
    
    if(args.d==None):
        args.d=time.strftime('%Y%m%d',time.localtime())
        
    if(args.i!=None):
        working_dir=os.getcwd()
        infile=os.path.join(working_dir,args.i)        
        if (os.path.exists(infile)):
            convert(infile,args.d)
        else:
            print "No such directory or file exsits"
